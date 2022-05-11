import io
import logging

import openpyxl


def save_workbook_to_file(workbook, filename):
    return workbook.save(filename)


def load_workbook_from_bytes(file_in_bytes):
    xlsx = io.BytesIO(file_in_bytes)
    return openpyxl.load_workbook(xlsx)


def load_and_save_workbook(file_name, file_in_bytes):
    wb = load_workbook_from_bytes(file_in_bytes)
    save_workbook_to_file(wb, file_name)
    logging.info(f'Файл {file_name} успешно сохранен.')
